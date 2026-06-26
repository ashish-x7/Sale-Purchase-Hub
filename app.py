import os
import io
import time
import math
import json
import traceback
import pickle
import tempfile
import requests
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 250 * 1024 * 1024  # 250MB max upload

# ─── Persistent Data Storage ────────────────────────────────────────────────
# Store data in the workspace folder so it persists between restarts
WORKSPACE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.environ.get('DATA_DIR')
if not DATA_DIR:
    DATA_DIR = os.path.join(tempfile.gettempdir(), 'sale_purchase_hub') if os.environ.get('RENDER') else WORKSPACE

UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
DATA_STORE_PATH = os.path.join(DATA_DIR, 'processed_data_store.pkl')


@app.errorhandler(413)
def request_too_large(error):
    return jsonify({'error': 'Uploaded files are too large. Please reduce file size or upload smaller files.'}), 413


@app.errorhandler(500)
def internal_server_error(error):
    traceback.print_exc()
    return jsonify({'error': 'Server error while processing. Please try again; if it repeats, check Render logs.'}), 500



def get_financial_year():
    """Get current financial year string like '2026-27'."""
    now = datetime.now()
    if now.month >= 4:  # April onwards = new FY
        return f'{now.year}-{str(now.year + 1)[2:]}'
    else:
        return f'{now.year - 1}-{str(now.year)[2:]}'


def load_stored_data():
    """Load previously stored data from disk."""
    if os.path.exists(DATA_STORE_PATH):
        try:
            with open(DATA_STORE_PATH, 'rb') as f:
                return normalize_stored_data(pickle.load(f))
        except Exception:
            pass
    return {'sale': [], 'purchase': [], 'sale_ids': set(), 'purchase_ids': set()}


def save_stored_data(data):
    """Save data to disk for persistence."""
    with open(DATA_STORE_PATH, 'wb') as f:
        pickle.dump(normalize_stored_data(data), f)


def clean_json_value(value):
    """Return a value that can be safely serialized as strict JSON."""
    if isinstance(value, float):
        return value if math.isfinite(value) else 0
    if isinstance(value, dict):
        return {k: clean_json_value(v) for k, v in value.items()}
    if isinstance(value, list):
        return [clean_json_value(v) for v in value]
    return value


def clean_rows(rows):
    return [clean_json_value(dict(row)) for row in rows]


def normalize_stored_data(data):
    """Normalize old/new pickle data into the current strict JSON-safe shape."""
    if not isinstance(data, dict):
        return {'sale': [], 'purchase': [], 'sale_ids': set(), 'purchase_ids': set()}

    sale = clean_rows(data.get('sale', []))
    purchase = clean_rows(data.get('purchase', []))
    sale_ids = set(data.get('sale_ids') or [])
    purchase_ids = set(data.get('purchase_ids') or [])

    if not sale_ids:
        sale_ids = {row.get('sale_unique_id') for row in sale if row.get('sale_unique_id')}
    if not purchase_ids:
        purchase_ids = {row.get('purchase_unique_id') for row in purchase if row.get('purchase_unique_id')}

    return {
        'sale': sale,
        'purchase': purchase,
        'sale_ids': sale_ids,
        'purchase_ids': purchase_ids,
    }


def merge_with_dedup(existing_data, new_data, existing_ids, unique_id_key):
    """Merge new data into existing, skipping duplicates by UNIQUE ID.
    
    Returns: (merged_data, merged_ids, new_count, duplicate_count)
    """
    new_count = 0
    dup_count = 0
    merged = list(existing_data)
    merged_ids = set(existing_ids)
    
    for row in new_data:
        uid = row.get(unique_id_key, '')
        if uid and uid in merged_ids:
            dup_count += 1
        else:
            merged.append(row)
            if uid:
                merged_ids.add(uid)
            new_count += 1
    
    # Re-number all rows
    for i, row in enumerate(merged):
        row['no'] = i + 1
    
    return merged, merged_ids, new_count, dup_count

# ─── Helpers ──────────────────────────────────────────────────────────

def roundup2(value):
    """Round up to 2 decimal places (like Excel ROUNDUP)."""
    if value is None or value == '':
        return 0
    try:
        v = float(value)
        if not math.isfinite(v):
            return 0
        return math.ceil(v * 100) / 100
    except (ValueError, TypeError):
        return 0


def safe_float(value, default=0):
    """Safely convert to float."""
    if value is None or value == '':
        return default
    try:
        v = float(value)
        return v if math.isfinite(v) else default
    except (ValueError, TypeError):
        return default


def safe_str(value, default=''):
    """Safely convert to string."""
    if value is None:
        return default
    s = str(value).strip()
    # Remove trailing .0 from numeric strings
    if s.endswith('.0'):
        try:
            int_val = int(float(s))
            if float(s) == int_val:
                return str(int_val)
        except (ValueError, TypeError):
            pass
    return s


def extract_state_short(state_code_full):
    """Extract short state code: 'Gujarat(GJ)' → '(GJ)'"""
    if not state_code_full:
        return ''
    s = str(state_code_full)
    # Find content in parentheses
    start = s.find('(')
    if start != -1:
        end = s.find(')', start)
        if end != -1:
            return s[start:end + 1]
    return s


def read_xls_file(filepath):
    """Read .xls file handling corruption."""
    try:
        wb = xlrd.open_workbook(filepath, ignore_workbook_corruption=True)
        ws = wb.sheet_by_index(0)
        
        # Read all data into list of dicts
        headers = []
        seen = {}
        for c in range(ws.ncols):
            h = str(ws.cell_value(1, c)).strip()  # Row 2 = headers (index 1)
            if not h or h == '':
                h = f'Col_{c}'
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 1
            headers.append(h)
        
        rows = []
        for r in range(2, ws.nrows):  # Skip title row and header row
            row_data = {}
            for c in range(ws.ncols):
                cell_value = ws.cell_value(r, c)
                # Handle dates
                if ws.cell_type(r, c) == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, wb.datemode)
                        cell_value = f'{date_tuple[2]:02d}/{date_tuple[1]:02d}/{date_tuple[0]}'
                    except Exception:
                        pass
                row_data[headers[c]] = cell_value
            rows.append(row_data)
        
        return headers, rows
    except Exception as e:
        raise Exception(f'Error reading {os.path.basename(filepath)}: {str(e)}')


def read_xlsx_file(filepath):
    """Read .xlsx file."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        headers = []
        seen = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=2, column=c).value  # Row 2 = headers
            if h is None:
                h = f'Col_{c}'
            h = str(h).strip()
            if h in seen:
                seen[h] += 1
                h = f"{h}_{seen[h]}"
            else:
                seen[h] = 1
            headers.append(h)
        
        rows = []
        for r in range(3, ws.max_row + 1):  # Data starts at row 3
            row_data = {}
            for c in range(1, ws.max_column + 1):
                row_data[headers[c - 1]] = ws.cell(row=r, column=c).value
            rows.append(row_data)
        
        return headers, rows
    except Exception as e:
        raise Exception(f'Error reading {os.path.basename(filepath)}: {str(e)}')


def read_file(filepath):
    """Read Excel file (auto-detect format)."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return read_xls_file(filepath)
    elif ext == '.xlsx':
        return read_xlsx_file(filepath)
    else:
        raise Exception(f'Unsupported file format: {ext}')


# ─── LOOKUP Logic ────────────────────────────────────────────────────────

def build_sale_summary_lookup(sale_summary_rows, sale_summary_headers):
    """Build lookup dict from SALE SUMMARY keyed by New Invoice No."""
    lookup = {}
    for row in sale_summary_rows:
        # Try "New Invoice No." first, then "New Invoice No"
        key = None
        for possible_key in ['New Invoice No.', 'New Invoice No', 'New invoice No.']:
            if possible_key in row:
                key = safe_str(row[possible_key])
                break
        
        if key:
            lookup[key] = {
                'invoice_type': safe_str(row.get('Invoice Type', '')),
                'invoice_id': safe_str(row.get('Invoice No', row.get('Invoice No.', ''))),
                'state_code': safe_str(row.get('State Code', '')),
                'channel': safe_str(row.get('Channel', '')),
            }
    return lookup


def build_purchase_summary_lookup(purchase_summary_rows, purchase_summary_headers):
    """Build lookup dict from PURCHASE SUMMARY keyed by Invoice No."""
    lookup = {}
    for row in purchase_summary_rows:
        key = None
        for possible_key in ['Invoice No.', 'Invoice No', 'Invoice Number']:
            if possible_key in row:
                key = safe_str(row[possible_key])
                break
        
        if key:
            lookup[key] = {
                'invoice_type': safe_str(row.get('Invoice Type', '')),
                'order_no': safe_str(row.get('Order No', '')),
            }
    return lookup


# ─── Main Processing ───────────────────────────────────────────────────────

def process_files(sale_details_path, sale_summary_path, purchase_details_path, purchase_summary_path):
    """Process 4 files and generate combined data."""
    
    # Read all files
    sd_headers, sd_rows = read_file(sale_details_path)
    ss_headers, ss_rows = read_file(sale_summary_path)
    pd_headers, pd_rows = read_file(purchase_details_path)
    ps_headers, ps_rows = read_file(purchase_summary_path)
    
    # Build LOOKUP dicts
    sale_lookup = build_sale_summary_lookup(ss_rows, ss_headers)
    purchase_lookup = build_purchase_summary_lookup(ps_rows, ps_headers)
    
    # ─── Process SALE side ──────────────────────────────────────────────
    sale_processed = []
    for idx, row in enumerate(sd_rows):
        invoice_no = safe_str(row.get('Invoice No', ''))
        
        # VLOOKUP from SALE SUMMARY
        summary = sale_lookup.get(invoice_no, {})
        
        # Direct columns from SALE DETAILS
        quantity = safe_float(row.get('Quantity', 0))
        item_cost = safe_float(row.get('Item Cost', row.get('Item cost', 0)))
        gross = safe_float(row.get('Gross', 0))
        igst_rate = safe_float(row.get('IGST', 0))
        cgst_rate = safe_float(row.get('CGST', 0))
        sgst_rate = safe_float(row.get('SGST', 0))
        igst_amt = safe_float(row.get('IGST Amt', 0))
        cgst_amt = safe_float(row.get('CGST Amt', 0))
        sgst_amt = safe_float(row.get('SGST Amt', 0))
        invoice_val = safe_float(row.get('Invoice', 0))
        
        order_id = safe_str(row.get('Order ID', ''))
        item_asin = safe_str(row.get('Item Asin', ''))
        item_sku = safe_str(row.get('Item SKU', ''))
        
        # SALE UNIQUE ID = InvoiceNo - OrderID - ItemAsin - ItemSKU
        sale_unique_id = f'{invoice_no}-{order_id}-{item_asin}-{item_sku}'
        
        # State code
        state_code_full = summary.get('state_code', '')
        state_code_short = extract_state_short(state_code_full)
        
        # Duplicate calculated columns (Col 28-37)
        calc_qty = quantity  # Col 28 = Col 13
        calc_cost = item_cost  # Col 29 = Col 14
        calc_gross = roundup2(calc_qty * calc_cost)  # Col 30 = 28 * 29
        calc_igst = igst_rate  # Col 31 = Col 16
        calc_cgst = cgst_rate  # Col 32 = Col 17
        calc_sgst = sgst_rate  # Col 33 = Col 18
        calc_igst_amt = roundup2(calc_gross * calc_igst / 100) if calc_igst else 0  # Col 34
        calc_cgst_amt = roundup2(calc_gross * calc_cgst / 100) if calc_cgst else 0  # Col 35
        calc_sgst_amt = roundup2(calc_gross * calc_sgst / 100) if calc_sgst else 0  # Col 36
        calc_invoice = calc_gross + calc_igst + calc_cgst + calc_sgst  # Col 37
        
        sale_row = {
            'no': idx + 1,
            'invoice_no': invoice_no,
            'type': summary.get('invoice_type', ''),
            'invoice_date': safe_str(row.get('Invoice Date', '')),
            'warehouse_name': safe_str(row.get('Warehouse Name', '')),
            'warehouse_code': safe_str(row.get('Warehouse Code', '')),
            'gst_no': safe_str(row.get('GST No', '')),
            'order_id': order_id,
            'item_asin': item_asin,
            'item_sku': item_sku,
            'item_name': safe_str(row.get('Item Name', '')),
            'hsn_number': safe_str(row.get('HSN Number', '')),
            'quantity': quantity,
            'item_cost': item_cost,
            'gross': gross,
            'igst': igst_rate,
            'cgst': cgst_rate,
            'sgst': sgst_rate,
            'igst_amt': igst_amt,
            'cgst_amt': cgst_amt,
            'sgst_amt': sgst_amt,
            'invoice': invoice_val,
            'reason': safe_str(row.get('Reason', '')),
            'zoho_status': safe_str(row.get('Zoho Status', '')),
            'invoice_id': summary.get('invoice_id', ''),
            'state_code': state_code_full,
            'sale_unique_id': sale_unique_id,
            # Duplicate calculated columns
            'calc_qty': calc_qty,
            'calc_cost': calc_cost,
            'calc_gross': calc_gross,
            'calc_igst': calc_igst,
            'calc_cgst': calc_cgst,
            'calc_sgst': calc_sgst,
            'calc_igst_amt': calc_igst_amt,
            'calc_cgst_amt': calc_cgst_amt,
            'calc_sgst_amt': calc_sgst_amt,
            'calc_invoice': roundup2(calc_invoice),
            'state_code_short': state_code_short,
        }
        sale_processed.append(sale_row)
    
    # ─── Process PURCHASE side ──────────────────────────────────────────
    purchase_processed = []
    for idx, row in enumerate(pd_rows):
        invoice_no = safe_str(row.get('Invoice No', ''))
        
        order_id = safe_str(row.get('Order ID', ''))
        item_asin = safe_str(row.get('Item Asin', ''))
        item_sku = safe_str(row.get('Item SKU', ''))
        
        quantity = safe_float(row.get('Quantity', 0))
        
        # Handle the duplicate "Quantity" column that's actually "Item cost"
        # In PURCHASE DETAILS, Col 12 header says "Quantity" but value is Item cost
        item_cost = 0
        if 'Quantity_2' in row:
            item_cost = safe_float(row.get('Quantity_2', 0))
        else:
            item_cost_key = 'Item cost'
            if item_cost_key not in row:
                item_cost_key = 'Item Cost'
            item_cost = safe_float(row.get(item_cost_key, 0))
            if item_cost == 0:
                # Fallback: the item cost might be mislabeled
                # Look at all values and find the one that looks like a cost
                for k, v in row.items():
                    if k not in [item_cost_key] and 'cost' in k.lower():
                        item_cost = safe_float(v)
                        break
        
        gross = safe_float(row.get('Gross', 0))
        igst_rate = safe_float(row.get('IGST', 0))
        cgst_rate = safe_float(row.get('CGST', 0))
        sgst_rate = safe_float(row.get('SGST', 0))
        igst_amt = safe_float(row.get('IGST Amt', 0))
        cgst_amt = safe_float(row.get('CGST Amt', 0))
        sgst_amt = safe_float(row.get('SGST Amt', 0))
        invoice_val = safe_float(row.get('Invoice', 0))
        
        # PURCHASE UNIQUE ID = InvoiceNo - OrderID - ItemAsin - ItemSKU (with "-" separator)
        purchase_unique_id = f'{invoice_no}-{order_id}-{item_asin}-{item_sku}'
        
        # Duplicate calculated columns (Col 61-69)
        calc_qty = quantity  # Col 61
        calc_cost = item_cost  # Col 62
        calc_gross = roundup2(calc_qty * calc_cost)  # Col 63
        calc_igst = igst_rate  # Col 64
        calc_cgst = cgst_rate  # Col 65
        calc_sgst = sgst_rate  # Col 66
        calc_igst_amt = roundup2(calc_gross * calc_igst / 100) if calc_igst else 0  # Col 67
        calc_cgst_amt = roundup2(calc_gross * calc_cgst / 100) if calc_cgst else 0  # Col 68
        calc_sgst_amt = roundup2(calc_gross * calc_sgst / 100) if calc_sgst else 0  # Col 69
        calc_total_amt = roundup2(calc_gross + calc_igst_amt + calc_cgst_amt + calc_sgst_amt)  # Col 70
        
        purchase_row = {
            'no': idx + 1,
            'invoice_no': invoice_no,
            'warehouse_name': safe_str(row.get('Warehouse Name', '')),
            'warehouse_code': safe_str(row.get('Warehouse Code', '')),
            'gst_no': safe_str(row.get('GST No', '')),
            'order_id': order_id,
            'item_asin': item_asin,
            'item_sku': item_sku,
            'item_name': safe_str(row.get('Item Name', '')),
            'hsn_number': safe_str(row.get('HSN Number', '')),
            'quantity': quantity,
            'item_cost': item_cost,
            'gross': gross,
            'igst': igst_rate,
            'cgst': cgst_rate,
            'sgst': sgst_rate,
            'igst_amt': igst_amt,
            'cgst_amt': cgst_amt,
            'sgst_amt': sgst_amt,
            'invoice': invoice_val,
            'purchase_unique_id': purchase_unique_id,
            # Duplicate calculated columns
            'calc_qty': calc_qty,
            'calc_cost': calc_cost,
            'calc_gross': calc_gross,
            'calc_igst': calc_igst,
            'calc_cgst': calc_cgst,
            'calc_sgst': calc_sgst,
            'calc_igst_amt': calc_igst_amt,
            'calc_cgst_amt': calc_cgst_amt,
            'calc_sgst_amt': calc_sgst_amt,
            'calc_total_amt': calc_total_amt,
        }
        purchase_processed.append(purchase_row)
    
    return sale_processed, purchase_processed


# ─── Excel Export ────────────────────────────────────────────────────────

def export_to_excel(sale_data, purchase_data, output_path):
    """Export combined data to Excel in the final format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    fy = get_financial_year()
    ws.title = f'AJIO & MYNTRA SALE-PURCHASE {fy}'
    
    # ─── Styles ─────────────────────────────────────────────────────────
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    sale_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    purchase_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    subheader_font = Font(name='Calibri', bold=True, size=10)
    sale_subheader_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    purchase_subheader_fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # ─── Row 1: Merged Headers ──────────────────────────────────────────
    ws.merge_cells('A1:AL1')  # SALE
    ws.merge_cells('AN1:BR1')  # PURCHASE
    
    sale_cell = ws.cell(row=1, column=1, value='SALE')
    sale_cell.font = header_font
    sale_cell.fill = sale_fill
    sale_cell.alignment = center_align
    
    purchase_cell = ws.cell(row=1, column=40, value='PURCHASE')
    purchase_cell.font = header_font
    purchase_cell.fill = purchase_fill
    purchase_cell.alignment = center_align
    
    # ─── Row 2: Column Headers ──────────────────────────────────────────
    sale_headers = [
        'No.', 'Invoice No', 'TYPE', 'Invoice Date', 'Warehouse Name',
        'Warehouse Code', 'GST No', 'Order ID', 'Item Asin', 'Item SKU',
        'Item Name', 'HSN Number', 'Quantity', 'Item Cost', 'Gross',
        'IGST', 'CGST', 'SGST', 'IGST Amt', 'CGST Amt', 'SGST Amt',
        'Invoice', 'Reason', 'Zoho Status', 'invoice id', 'State Code',
        'SALE UNIQUE ID', 'Quantity', 'Item Cost', 'Gross',
        'IGST', 'CGST', 'SGST', 'IGST Amt', 'CGST Amt', 'SGST Amt',
        'Invoice', 'STATE CODE'
    ]
    
    purchase_headers = [
        'No.', 'Invoice No', 'Warehouse Name', 'Warehouse Code', 'GST No',
        'Order ID', 'Item Asin', 'Item SKU', 'Item Name', 'HSN Number',
        'Quantity', 'Item cost', 'Gross', 'IGST', 'CGST', 'SGST',
        'IGST Amt', 'CGST Amt', 'SGST Amt', 'Invoice',
        'PURCHASE UNIQUE ID', 'Quantity', 'Item cost', 'Gross',
        'IGST', 'CGST', 'SGST', 'IGST Amt', 'CGST Amt', 'SGST Amt',
        'Total Amt'
    ]
    
    # Write SALE headers (Col 1-38)
    for i, h in enumerate(sale_headers):
        cell = ws.cell(row=2, column=i + 1, value=h)
        cell.font = subheader_font
        cell.fill = sale_subheader_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Write PURCHASE headers (Col 40-70)
    for i, h in enumerate(purchase_headers):
        cell = ws.cell(row=2, column=40 + i, value=h)
        cell.font = subheader_font
        cell.fill = purchase_subheader_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # ─── Data Rows ──────────────────────────────────────────────────────
    max_rows = max(len(sale_data), len(purchase_data))
    
    sale_keys = [
        'no', 'invoice_no', 'type', 'invoice_date', 'warehouse_name',
        'warehouse_code', 'gst_no', 'order_id', 'item_asin', 'item_sku',
        'item_name', 'hsn_number', 'quantity', 'item_cost', 'gross',
        'igst', 'cgst', 'sgst', 'igst_amt', 'cgst_amt', 'sgst_amt',
        'invoice', 'reason', 'zoho_status', 'invoice_id', 'state_code',
        'sale_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
        'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
        'calc_cgst_amt', 'calc_sgst_amt', 'calc_invoice', 'state_code_short'
    ]
    
    purchase_keys = [
        'no', 'invoice_no', 'warehouse_name', 'warehouse_code', 'gst_no',
        'order_id', 'item_asin', 'item_sku', 'item_name', 'hsn_number',
        'quantity', 'item_cost', 'gross', 'igst', 'cgst', 'sgst',
        'igst_amt', 'cgst_amt', 'sgst_amt', 'invoice',
        'purchase_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
        'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
        'calc_cgst_amt', 'calc_sgst_amt', 'calc_total_amt'
    ]
    
    for r in range(max_rows):
        row_num = r + 3  # Data starts at row 3
        
        # SALE data
        if r < len(sale_data):
            for c, key in enumerate(sale_keys):
                cell = ws.cell(row=row_num, column=c + 1, value=sale_data[r].get(key, ''))
                cell.border = thin_border
        
        # PURCHASE data
        if r < len(purchase_data):
            for c, key in enumerate(purchase_keys):
                cell = ws.cell(row=row_num, column=40 + c, value=purchase_data[r].get(key, ''))
                cell.border = thin_border
    
    # ─── Column Widths ──────────────────────────────────────────────────
    for col in range(1, 72):
        ws.column_dimensions[get_column_letter(col)].width = 15
    # Wider columns for names
    ws.column_dimensions['K'].width = 40  # Item Name
    ws.column_dimensions['AW'].width = 40  # Purchase Item Name
    
    wb.save(output_path)
    return output_path


# ─── Routes ──────────────────────────────────────────────────────────

# Cache storage for Google Sheet Sale-Purchase Quantities (SQLite-based for low memory)
import threading
import re
import csv
import sqlite3
import gc

_CACHE_DB_PATH = os.path.join(DATA_DIR, 'sale_cache.db')
_cache_loaded_event = threading.Event()
_is_cache_loading = False

def _init_cache_db():
    """Initialize the SQLite cache database."""
    # Check if we need to migrate the database
    if os.path.exists(_CACHE_DB_PATH):
        try:
            conn = sqlite3.connect(_CACHE_DB_PATH)
            cursor = conn.cursor()
            cursor.execute("PRAGMA table_info(sale_cache)")
            columns = [col[1] for col in cursor.fetchall()]
            conn.close()
            # Migrate if any new column like 'invoice_no' or 'igst_amt' is missing
            if columns and ('invoice_no' not in columns or 'igst_amt' not in columns):
                print("[CACHE SYNC] Outdated SQLite schema detected. Deleting old DB to trigger fresh load.", flush=True)
                try:
                    os.remove(_CACHE_DB_PATH)
                except Exception as del_err:
                    print(f"[CACHE SYNC] Failed to delete old DB: {str(del_err)}", flush=True)
        except Exception as check_err:
            print(f"[CACHE SYNC] Error checking schema migration: {str(check_err)}", flush=True)

    conn = sqlite3.connect(_CACHE_DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sale_cache (
            key TEXT PRIMARY KEY, qty REAL, rate REAL, state TEXT, seller_name TEXT, sheet_name TEXT,
            invoice_no TEXT, invoice_date TEXT, warehouse_code TEXT, gst_no TEXT,
            order_id TEXT, item_asin TEXT, item_sku TEXT, item_name TEXT, hsn_number TEXT,
            invoice_val REAL, reason TEXT, zoho_status TEXT, invoice_id TEXT,
            gross REAL, igst REAL, cgst REAL, sgst REAL, igst_amt REAL, cgst_amt REAL, sgst_amt REAL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_key ON sale_cache(key)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS purchase_cache (
            key TEXT PRIMARY KEY, qty REAL, rate REAL, seller_name TEXT, sheet_name TEXT,
            invoice_no TEXT, warehouse_code TEXT, gst_no TEXT, order_id TEXT,
            item_asin TEXT, item_sku TEXT, item_name TEXT, hsn_number TEXT,
            invoice_val REAL,
            gross REAL, igst REAL, cgst REAL, sgst REAL, igst_amt REAL, cgst_amt REAL, sgst_amt REAL
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_key_purchase ON purchase_cache(key)")
    conn.commit()
    conn.close()

print("[DATABASE PATH] Resolving _CACHE_DB_PATH:", _CACHE_DB_PATH, flush=True)
_init_cache_db()

def _fetch_and_store_google_sheet_to_sqlite():
    """Download Google Sheet CSV data and store directly into SQLite (low memory)."""
    spreadsheet_id = "1xXR4gfDPN-0A1B8gaY7kcW52gr4uRjfJ-TmafuJr6To"
    edit_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    
    try:
        res = requests.get(edit_url, timeout=30)
        if res.status_code != 200:
            print(f"[CACHE SYNC ERROR] Failed to fetch edit page. Status: {res.status_code}", flush=True)
            return False
            
        html = res.text
        match = re.search(r'bootstrapData\s*=\s*({.+?});', html)
        if not match:
            match = re.search(r'var\s+bootstrapData\s*=\s*({.+?});', html)
            
        sheets = []
        if match:
            try:
                data = json.loads(match.group(1))
                topsnapshot = data.get("changes", {}).get("topsnapshot", [])
                for item in topsnapshot:
                    if not isinstance(item, list) or len(item) < 2:
                        continue
                    payload_str = item[1]
                    if isinstance(payload_str, str) and payload_str.startswith("["):
                        inner = json.loads(payload_str)
                        if isinstance(inner, list) and len(inner) >= 4:
                            gid = str(inner[2])
                            props_list = inner[3]
                            if isinstance(props_list, list) and len(props_list) > 0:
                                props_dict = props_list[0]
                                if isinstance(props_dict, dict) and "1" in props_dict:
                                    title_block = props_dict["1"]
                                    if isinstance(title_block, list) and len(title_block) > 0:
                                        title_info = title_block[0]
                                        if isinstance(title_info, list) and len(title_info) >= 3:
                                            title = title_info[2]
                                            sheets.append((gid, title))
            except Exception as e:
                print(f"[CACHE SYNC ERROR] Parsing bootstrapData failed: {str(e)}", flush=True)
        
        # Free the HTML string immediately
        del html, res
        gc.collect()
                
        if not sheets:
            sheets = [("0", "AJ&MY&FK SALE-PURCHASE 2025-26"), ("1563945167", "AJ&MY&FK SALE-PURCHASE 2026-27")]
        
        # Use a temporary table in the main database
        conn = sqlite3.connect(_CACHE_DB_PATH)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("DROP TABLE IF EXISTS sale_cache_new")
        conn.execute("""
            CREATE TABLE sale_cache_new (
                key TEXT PRIMARY KEY, qty REAL, rate REAL, state TEXT, seller_name TEXT, sheet_name TEXT,
                invoice_no TEXT, invoice_date TEXT, warehouse_code TEXT, gst_no TEXT,
                order_id TEXT, item_asin TEXT, item_sku TEXT, item_name TEXT, hsn_number TEXT,
                invoice_val REAL, reason TEXT, zoho_status TEXT, invoice_id TEXT,
                gross REAL, igst REAL, cgst REAL, sgst REAL, igst_amt REAL, cgst_amt REAL, sgst_amt REAL
            )
        """)
        conn.execute("DROP TABLE IF EXISTS purchase_cache_new")
        conn.execute("""
            CREATE TABLE purchase_cache_new (
                key TEXT PRIMARY KEY, qty REAL, rate REAL, seller_name TEXT, sheet_name TEXT,
                invoice_no TEXT, warehouse_code TEXT, gst_no TEXT, order_id TEXT,
                item_asin TEXT, item_sku TEXT, item_name TEXT, hsn_number TEXT,
                invoice_val REAL,
                gross REAL, igst REAL, cgst REAL, sgst REAL, igst_amt REAL, cgst_amt REAL, sgst_amt REAL
            )
        """)
        conn.commit()
        
        total_keys = 0
        for gid, title in sheets:
            title_upper = title.upper()
            if any(x in title_upper for x in ["DASHBOARD", "SUMMARY", "CONFIG", "INSTRUCTION", "LIMIT", "USER", "PART", "REPORT", "WELCOME"]):
                continue
                
            csv_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=csv&gid={gid}"
            print(f"[CACHE SYNC] Downloading CSV for tab '{title}' (GID {gid})...", flush=True)
            
            try:
                # Stream the CSV to avoid loading entire response into memory
                res_csv = requests.get(csv_url, timeout=60, stream=True)
                if res_csv.status_code == 200:
                    # Decode streaming content line by line
                    lines = (line.decode('utf-8') for line in res_csv.iter_lines() if line)
                    reader = csv.reader(lines)
                    
                    sheet_count = 0
                    batch = []
                    pur_batch = []
                    for r_idx, row in enumerate(reader):
                        if r_idx < 2:  # Skip header rows
                            continue
                        if len(row) < 28:
                            continue
                            
                        key_val = row[26].strip()
                        qty_val = row[27].strip()
                        rate_val = row[28].strip() if len(row) > 28 else ""
                        state_val = row[37].strip() if len(row) > 37 else ""
                        seller_name_val = row[4].strip() if len(row) > 4 else ""
                        
                        # Fallback key construction
                        if not key_val:
                            inv = row[1].strip()
                            ord_id = row[7].strip()
                            asin = row[8].strip()
                            sku = row[9].strip()
                            if inv and (ord_id or sku):
                                key_val = f"{inv}-{ord_id}-{asin}-{sku}"
                                
                        if key_val:
                            norm_key = key_val.strip().upper()
                            try:
                                qty = float(qty_val) if qty_val else 0
                            except ValueError:
                                qty = 0
                            try:
                                rate = float(rate_val) if rate_val else 0
                            except ValueError:
                                rate = 0
                            state = state_val.strip()
                            
                            # Additional sale fields
                            invoice_no = row[1].strip() if len(row) > 1 else ""
                            invoice_date = row[3].strip() if len(row) > 3 else ""
                            warehouse_code = row[5].strip() if len(row) > 5 else ""
                            gst_no = row[6].strip() if len(row) > 6 else ""
                            order_id = row[7].strip() if len(row) > 7 else ""
                            item_asin = row[8].strip() if len(row) > 8 else ""
                            item_sku = row[9].strip() if len(row) > 9 else ""
                            item_name = row[10].strip() if len(row) > 10 else ""
                            hsn_number = row[11].strip() if len(row) > 11 else ""
                            try:
                                invoice_val = float(row[21].strip()) if len(row) > 21 and row[21].strip() else 0.0
                            except ValueError:
                                invoice_val = 0.0
                            reason = row[22].strip() if len(row) > 22 else ""
                            zoho_status = row[23].strip() if len(row) > 23 else ""
                            invoice_id = row[24].strip() if len(row) > 24 else ""
                            
                            try: gross = float(row[14].strip()) if len(row) > 14 and row[14].strip() else 0.0
                            except: gross = 0.0
                            try: igst = float(row[15].strip()) if len(row) > 15 and row[15].strip() else 0.0
                            except: igst = 0.0
                            try: cgst = float(row[16].strip()) if len(row) > 16 and row[16].strip() else 0.0
                            except: cgst = 0.0
                            try: sgst = float(row[17].strip()) if len(row) > 17 and row[17].strip() else 0.0
                            except: sgst = 0.0
                            try: igst_amt = float(row[18].strip()) if len(row) > 18 and row[18].strip() else 0.0
                            except: igst_amt = 0.0
                            try: cgst_amt = float(row[19].strip()) if len(row) > 19 and row[19].strip() else 0.0
                            except: cgst_amt = 0.0
                            try: sgst_amt = float(row[20].strip()) if len(row) > 20 and row[20].strip() else 0.0
                            except: sgst_amt = 0.0
                            
                            batch.append((
                                norm_key, qty, rate, state, seller_name_val, title,
                                invoice_no, invoice_date, warehouse_code, gst_no,
                                order_id, item_asin, item_sku, item_name, hsn_number,
                                invoice_val, reason, zoho_status, invoice_id,
                                gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
                            ))
                            sheet_count += 1
                            
                            if len(batch) >= 5000:
                                conn.executemany("""
                                    INSERT OR REPLACE INTO sale_cache_new (
                                        key, qty, rate, state, seller_name, sheet_name,
                                        invoice_no, invoice_date, warehouse_code, gst_no,
                                        order_id, item_asin, item_sku, item_name, hsn_number,
                                        invoice_val, reason, zoho_status, invoice_id,
                                        gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, batch)
                                conn.commit()
                                batch = []
                                
                        pur_key_val = row[59].strip() if len(row) > 59 else ""
                        if pur_key_val:
                            pur_norm_key = pur_key_val.strip().upper()
                            try:
                                pur_qty = float(row[60].strip()) if len(row) > 60 and row[60].strip() else 0
                            except ValueError:
                                pur_qty = 0
                            try:
                                pur_rate = float(row[61].strip()) if len(row) > 61 and row[61].strip() else 0
                            except ValueError:
                                pur_rate = 0
                            pur_seller_name_val = row[41].strip() if len(row) > 41 else ""
                            
                            # Additional purchase fields
                            pur_invoice_no = row[40].strip() if len(row) > 40 else ""
                            pur_warehouse_code = row[42].strip() if len(row) > 42 else ""
                            pur_gst_no = row[43].strip() if len(row) > 43 else ""
                            pur_order_id = row[44].strip() if len(row) > 44 else ""
                            pur_item_asin = row[45].strip() if len(row) > 45 else ""
                            pur_item_sku = row[46].strip() if len(row) > 46 else ""
                            pur_item_name = row[47].strip() if len(row) > 47 else ""
                            pur_hsn_number = row[48].strip() if len(row) > 48 else ""
                            try:
                                pur_invoice_val = float(row[58].strip()) if len(row) > 58 and row[58].strip() else 0.0
                            except ValueError:
                                pur_invoice_val = 0.0
                                
                            try: pur_gross = float(row[51].strip()) if len(row) > 51 and row[51].strip() else 0.0
                            except: pur_gross = 0.0
                            try: pur_igst = float(row[52].strip()) if len(row) > 52 and row[52].strip() else 0.0
                            except: pur_igst = 0.0
                            try: pur_cgst = float(row[53].strip()) if len(row) > 53 and row[53].strip() else 0.0
                            except: pur_cgst = 0.0
                            try: pur_sgst = float(row[54].strip()) if len(row) > 54 and row[54].strip() else 0.0
                            except: pur_sgst = 0.0
                            try: pur_igst_amt = float(row[55].strip()) if len(row) > 55 and row[55].strip() else 0.0
                            except: pur_igst_amt = 0.0
                            try: pur_cgst_amt = float(row[56].strip()) if len(row) > 56 and row[56].strip() else 0.0
                            except: pur_cgst_amt = 0.0
                            try: pur_sgst_amt = float(row[57].strip()) if len(row) > 57 and row[57].strip() else 0.0
                            except: pur_sgst_amt = 0.0
                                
                            pur_batch.append((
                                pur_norm_key, pur_qty, pur_rate, pur_seller_name_val, title,
                                pur_invoice_no, pur_warehouse_code, pur_gst_no, pur_order_id,
                                pur_item_asin, pur_item_sku, pur_item_name, pur_hsn_number,
                                pur_invoice_val,
                                pur_gross, pur_igst, pur_cgst, pur_sgst, pur_igst_amt, pur_cgst_amt, pur_sgst_amt
                            ))
                            
                            if len(pur_batch) >= 5000:
                                conn.executemany("""
                                    INSERT OR REPLACE INTO purchase_cache_new (
                                        key, qty, rate, seller_name, sheet_name,
                                        invoice_no, warehouse_code, gst_no, order_id,
                                        item_asin, item_sku, item_name, hsn_number,
                                        invoice_val,
                                        gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """, pur_batch)
                                conn.commit()
                                pur_batch = []
                    
                    # Insert remaining batch
                    if batch:
                        conn.executemany("""
                            INSERT OR REPLACE INTO sale_cache_new (
                                key, qty, rate, state, seller_name, sheet_name,
                                invoice_no, invoice_date, warehouse_code, gst_no,
                                order_id, item_asin, item_sku, item_name, hsn_number,
                                        invoice_val, reason, zoho_status, invoice_id,
                                        gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, batch)
                        conn.commit()
                        batch = []
                    if pur_batch:
                        conn.executemany("""
                            INSERT OR REPLACE INTO purchase_cache_new (
                                key, qty, rate, seller_name, sheet_name,
                                invoice_no, warehouse_code, gst_no, order_id,
                                        item_asin, item_sku, item_name, hsn_number,
                                        invoice_val,
                                        gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
                                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, pur_batch)
                        conn.commit()
                        pur_batch = []
                    
                    total_keys += sheet_count
                    print(f"[CACHE SYNC] Loaded {sheet_count} keys from tab '{title}'", flush=True)
                    
                    res_csv.close()
                    del res_csv
                    gc.collect()
                else:
                    print(f"[CACHE SYNC ERROR] Failed to fetch tab '{title}'. Status: {res_csv.status_code}", flush=True)
            except Exception as ex:
                print(f"[CACHE SYNC ERROR] Error loading tab '{title}': {str(ex)}", flush=True)
        
        # Create index on the temporary table
        conn.execute("CREATE INDEX IF NOT EXISTS idx_key_new ON sale_cache_new(key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_key_purchase_new ON purchase_cache_new(key)")
        conn.commit()
        
        # Atomically swap table names
        conn.execute("BEGIN TRANSACTION")
        try:
            conn.execute("DROP TABLE IF EXISTS sale_cache_old")
            conn.execute("DROP TABLE IF EXISTS purchase_cache_old")
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sale_cache'")
            if cursor.fetchone():
                conn.execute("ALTER TABLE sale_cache RENAME TO sale_cache_old")
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='purchase_cache'")
            if cursor.fetchone():
                conn.execute("ALTER TABLE purchase_cache RENAME TO purchase_cache_old")
            
            conn.execute("ALTER TABLE sale_cache_new RENAME TO sale_cache")
            conn.execute("ALTER TABLE purchase_cache_new RENAME TO purchase_cache")
            conn.execute("DROP TABLE IF EXISTS sale_cache_old")
            conn.execute("DROP TABLE IF EXISTS purchase_cache_old")
            conn.commit()
        except Exception as swap_ex:
            conn.rollback()
            raise swap_ex
        finally:
            conn.close()
        
        gc.collect()
        return total_keys > 0
    except Exception as e:
        print(f"[CACHE SYNC ERROR] Global sheet load failed: {str(e)}", flush=True)
        traceback.print_exc()
        try:
            conn.close()
        except:
            pass
        return False

def _load_cache_safely():
    global _is_cache_loading
    if _is_cache_loading:
        return
    try:
        _is_cache_loading = True
        
        # Check if database already exists on disk and has records
        has_existing_cache = False
        if os.path.exists(_CACHE_DB_PATH):
            try:
                conn = sqlite3.connect(_CACHE_DB_PATH)
                count = conn.execute("SELECT COUNT(*) FROM sale_cache").fetchone()[0]
                conn.close()
                if count > 0:
                    print(f"[CACHE SYNC] Found existing SQLite cache on disk with {count} keys. Marking loaded immediately.", flush=True)
                    _cache_loaded_event.set()
                    has_existing_cache = True
            except Exception as check_err:
                print(f"[CACHE SYNC] Error checking existing cache: {str(check_err)}", flush=True)
                
        print("[CACHE SYNC] Starting background sync from Google Sheets...", flush=True)
        success = _fetch_and_store_google_sheet_to_sqlite()
        if success:
            _cache_loaded_event.set()
            # Count keys in DB
            try:
                conn = sqlite3.connect(_CACHE_DB_PATH)
                count = conn.execute("SELECT COUNT(*) FROM sale_cache").fetchone()[0]
                conn.close()
                print(f"[CACHE SYNC SUCCESS] SQLite cache updated with {count} keys.", flush=True)
            except:
                print("[CACHE SYNC SUCCESS] SQLite cache updated.", flush=True)
        else:
            print("[CACHE SYNC WARNING] Failed to load data, keeping previous cache.", flush=True)
            if has_existing_cache:
                _cache_loaded_event.set()
    except Exception as e:
        print(f"[CACHE SYNC ERROR] {str(e)}", flush=True)
        traceback.print_exc()
        if has_existing_cache:
            _cache_loaded_event.set()
    finally:
        _is_cache_loading = False

def _background_cache_loader_loop():
    _load_cache_safely()
    while True:
        time.sleep(600)  # Refresh every 10 minutes
        _load_cache_safely()

_cache_loader_started = False
_cache_loader_lock = threading.Lock()

def start_background_cache_loader():
    global _cache_loader_started
    if _cache_loader_started:
        return
    with _cache_loader_lock:
        if _cache_loader_started:
            return
        print("[CACHE LOADER] Starting background cache loader thread...", flush=True)
        thread = threading.Thread(target=_background_cache_loader_loop, daemon=True)
        thread.start()
        _cache_loader_started = True

@app.route('/api/lookup-sale-quantities', methods=['POST', 'OPTIONS'])
def lookup_sale_quantities():
    if request.method == 'OPTIONS':
        response = jsonify({"status": "Success"})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "POST, OPTIONS")
        return response

    try:
        # Start background sync thread inside Gunicorn worker process lazily on first request
        start_background_cache_loader()
        
        print("[LOOKUP] Received lookup request!", flush=True)
        req_data = request.get_json(force=True, silent=True) or {}
        keys = req_data.get('keys', [])
        print(f"[LOOKUP] Keys count: {len(keys)}", flush=True)
        if not keys or not isinstance(keys, list):
            response = jsonify({"status": "Error", "message": "Invalid keys format"})
            response.headers.add("Access-Control-Allow-Origin", "*")
            return response, 400

        # Wait for the initial cache load to complete (up to 90 seconds if empty)
        if not _cache_loaded_event.is_set():
            print("[LOOKUP] Cache not loaded yet, waiting for initial sync...", flush=True)
            _cache_loaded_event.wait(timeout=90)
            if not _cache_loaded_event.is_set():
                print("[LOOKUP] Cache still not loaded after 90s!", flush=True)

        # Query SQLite directly — very low memory usage
        mapping = {}
        try:
            conn = sqlite3.connect(_CACHE_DB_PATH)
            # Query in batches to keep memory low
            norm_keys = [(str(k).strip().upper(),) for k in keys]
            # Use parameterized query with IN clause (batch of 500)
            for i in range(0, len(norm_keys), 500):
                batch = norm_keys[i:i+500]
                placeholders = ','.join(['?' for _ in batch])
                query = f"SELECT key, qty, rate, state, seller_name FROM sale_cache WHERE key IN ({placeholders})"
                results = conn.execute(query, [k[0] for k in batch]).fetchall()
                for db_key, db_qty, db_rate, db_state, db_seller_name in results:
                    # Find the original key (preserve case)
                    for orig_key in keys:
                        if str(orig_key).strip().upper() == db_key:
                            mapping[orig_key] = {
                                "qty": db_qty,
                                "rate": db_rate,
                                "state": db_state,
                                "sellerName": db_seller_name
                            }
                            break
            conn.close()
        except Exception as db_err:
            print(f"[LOOKUP] SQLite query error: {str(db_err)}", flush=True)

        # Merge with locally uploaded data from processed_data_store.pkl (if any)
        try:
            stored = load_stored_data()
            sale_rows = stored.get('sale', [])
            uploaded_mapping = {}
            for row in sale_rows:
                uid = row.get('sale_unique_id')
                if uid:
                    uid_str = str(uid).strip().upper()
                    uploaded_mapping[uid_str] = {
                        "qty": row.get('calc_qty', row.get('quantity', 0)),
                        "rate": row.get('calc_cost', row.get('item_cost', 0)),
                        "state": row.get('state_code_short', ''),
                        "sellerName": row.get('warehouse_name', '')
                    }
            # Check uploaded data for any keys not found in SQLite
            for key in keys:
                if key not in mapping:
                    norm_key = str(key).strip().upper()
                    if norm_key in uploaded_mapping:
                        mapping[key] = uploaded_mapping[norm_key]
        except Exception:
            pass

        print(f"[LOOKUP] Found {len(mapping)} matches out of {len(keys)} keys.", flush=True)
        response = jsonify({"status": "Success", "mapping": mapping})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response
    except Exception as e:
        traceback.print_exc()
        response = jsonify({"status": "Error", "message": str(e)})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response, 500

@app.route('/api/lookup-purchase-quantities', methods=['POST', 'OPTIONS'])
def lookup_purchase_quantities():
    if request.method == 'OPTIONS':
        response = jsonify({"status": "Success"})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "POST, OPTIONS")
        return response

    try:
        start_background_cache_loader()
        
        print("[LOOKUP PURCHASE] Received lookup request!", flush=True)
        req_data = request.get_json(force=True, silent=True) or {}
        keys = req_data.get('keys', [])
        print(f"[LOOKUP PURCHASE] Keys count: {len(keys)}", flush=True)
        if not keys or not isinstance(keys, list):
            response = jsonify({"status": "Error", "message": "Invalid keys format"})
            response.headers.add("Access-Control-Allow-Origin", "*")
            return response, 400

        if not _cache_loaded_event.is_set():
            print("[LOOKUP PURCHASE] Cache not loaded yet, waiting for initial sync...", flush=True)
            _cache_loaded_event.wait(timeout=90)

        mapping = {}
        try:
            conn = sqlite3.connect(_CACHE_DB_PATH)
            norm_keys = [(str(k).strip().upper(),) for k in keys]
            for i in range(0, len(norm_keys), 500):
                batch = norm_keys[i:i+500]
                placeholders = ','.join(['?' for _ in batch])
                query = f"SELECT key, qty, rate, state, seller_name FROM purchase_cache WHERE key IN ({placeholders})"
                results = conn.execute(query, [k[0] for k in batch]).fetchall()
                for db_key, db_qty, db_rate, db_state, db_seller_name in results:
                    for orig_key in keys:
                        if str(orig_key).strip().upper() == db_key:
                            mapping[orig_key] = {
                                "qty": db_qty,
                                "rate": db_rate,
                                "state": db_state,
                                "sellerName": db_seller_name
                            }
                            break
            conn.close()
        except Exception as db_err:
            print(f"[LOOKUP PURCHASE] SQLite query error: {str(db_err)}", flush=True)

        try:
            stored = load_stored_data()
            purchase_rows = stored.get('purchase', [])
            uploaded_mapping = {}
            for row in purchase_rows:
                uid = row.get('purchase_unique_id')
                if uid:
                    uid_str = str(uid).strip().upper()
                    uploaded_mapping[uid_str] = {
                        "qty": row.get('calc_qty', row.get('quantity', 0)),
                        "rate": row.get('calc_cost', row.get('item_cost', 0)),
                        "state": row.get('state_code_short', ''),
                        "sellerName": row.get('warehouse_name', '')
                    }
            for key in keys:
                if key not in mapping:
                    norm_key = str(key).strip().upper()
                    if norm_key in uploaded_mapping:
                        mapping[key] = uploaded_mapping[norm_key]
        except Exception:
            pass

        print(f"[LOOKUP PURCHASE] Found {len(mapping)} matches out of {len(keys)} keys.", flush=True)
        response = jsonify({"status": "Success", "mapping": mapping})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response
    except Exception as e:
        traceback.print_exc()
        response = jsonify({"status": "Error", "message": str(e)})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response, 500


@app.route('/api/get-available-sheets', methods=['GET', 'OPTIONS'])
def get_available_sheets():
    if request.method == 'OPTIONS':
        response = jsonify({"status": "Success"})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "GET, OPTIONS")
        return response

    try:
        # Lazy start of background loader
        start_background_cache_loader()
        
        # Wait up to 10s if cache is loading
        if not _cache_loaded_event.is_set():
            _cache_loaded_event.wait(timeout=10)

        conn = sqlite3.connect(_CACHE_DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT sheet_name FROM sale_cache WHERE sheet_name IS NOT NULL AND sheet_name != ''")
        sales_sheets = [r[0] for r in cursor.fetchall()]
        cursor.execute("SELECT DISTINCT sheet_name FROM purchase_cache WHERE sheet_name IS NOT NULL AND sheet_name != ''")
        purchases_sheets = [r[0] for r in cursor.fetchall()]
        conn.close()
        
        all_sheets = list(set(sales_sheets + purchases_sheets))
        
        import re
        years = []
        for s in all_sheets:
            m = re.search(r'\d{4}-\d{2}', s)
            if m:
                years.append(m.group(0))
            else:
                years.append(s)
                
        years = sorted(list(set(years)))
        if not years:
            years = ["2025-26", "2026-27"] # Fallbacks if empty
            
        response = jsonify({
            "status": "Success",
            "sheets": all_sheets,
            "years": years
        })
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response
    except Exception as e:
        traceback.print_exc()
        response = jsonify({"status": "Error", "message": str(e)})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response, 500


@app.route('/api/export-returns', methods=['POST', 'OPTIONS'])
def export_returns():
    if request.method == 'OPTIONS':
        response = jsonify({"status": "Success"})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add("Access-Control-Allow-Headers", "Content-Type")
        response.headers.add("Access-Control-Allow-Methods", "POST, OPTIONS")
        return response

    try:
        start_background_cache_loader()
        
        req_data = request.get_json(force=True, silent=True) or {}
        platform = req_data.get('platform', '').strip().upper()
        year = req_data.get('year', 'ALL').strip() # 'ALL' or '2025-26', '2026-27', etc.
        data_type = req_data.get('dataType', 'ALL').strip().upper() # 'ALL', 'SALE', 'PURCHASE'
        
        def matches_platform(plat, key, seller_name, row_type=""):
            if not plat or plat == 'ALL':
                return True
            plat = plat.upper()
            key = (key or '').upper()
            seller_name = (seller_name or '').upper()
            row_type = (row_type or '').upper()
            
            if plat == 'AJIO':
                # Exclude Myntra and Flipkart specific prefixes
                if key.startswith('MY') or key.startswith('FK'):
                    return False
                if key.startswith('AJ') or key.startswith('CG') or key.startswith('PG'):
                    return True
                if 'AJIO' in key or 'AJIO' in seller_name or 'AJ' in seller_name:
                    return True
            elif plat == 'MYNTRA':
                # Exclude Ajio and Flipkart specific prefixes
                if key.startswith('AJ') or key.startswith('FK'):
                    return False
                if key.startswith('MY') or key.startswith('CG') or key.startswith('PG'):
                    return True
                if 'MYNTRA' in key or 'MYNTRA' in seller_name or 'MY' in seller_name:
                    return True
            elif plat == 'FLIPKART':
                # Exclude Ajio and Myntra specific prefixes
                if key.startswith('AJ') or key.startswith('MY'):
                    return False
                if key.startswith('FK') or key.startswith('CG') or key.startswith('PG'):
                    return True
                if 'FLIPKART' in key or 'FLIPKART' in seller_name or 'FK' in seller_name:
                    return True
            return False

        if not _cache_loaded_event.is_set():
            _cache_loaded_event.wait(timeout=10)

        conn = sqlite3.connect(_CACHE_DB_PATH)
        
        # Build query for sales
        sale_query = """
            SELECT 
                key, qty, rate, state, seller_name, sheet_name,
                invoice_no, invoice_date, warehouse_code, gst_no,
                order_id, item_asin, item_sku, item_name, hsn_number,
                invoice_val, reason, zoho_status, invoice_id,
                gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
            FROM sale_cache
        """
        params = []
        conditions = []
        if year != 'ALL':
            conditions.append("sheet_name LIKE ?")
            params.append(f"%{year}%")
        if conditions:
            sale_query += " WHERE " + " AND ".join(conditions)
        
        sale_results = conn.execute(sale_query, params).fetchall()
        
        # Build query for purchases
        purchase_query = """
            SELECT 
                key, qty, rate, seller_name, sheet_name,
                invoice_no, warehouse_code, gst_no, order_id,
                item_asin, item_sku, item_name, hsn_number,
                invoice_val, gross, igst, cgst, sgst, igst_amt, cgst_amt, sgst_amt
            FROM purchase_cache
        """
        pur_params = []
        pur_conditions = []
        if year != 'ALL':
            pur_conditions.append("sheet_name LIKE ?")
            pur_params.append(f"%{year}%")
        if pur_conditions:
            purchase_query += " WHERE " + " AND ".join(pur_conditions)
            
        purchase_results = conn.execute(purchase_query, pur_params).fetchall()
        conn.close()
        
        # Format database records to match the expected Excel schema
        sale_rows = []
        for row in sale_results:
            db_key, db_qty, db_rate, db_state, db_seller_name, db_sheet = row[0:6]
            (
                invoice_no, invoice_date, warehouse_code, gst_no,
                order_id, item_asin, item_sku, item_name, hsn_number,
                invoice_val, reason, zoho_status, invoice_id,
                gross_val, igst_val, cgst_val, sgst_val, igst_amt_val, cgst_amt_val, sgst_amt_val
            ) = row[6:]
            
            if not matches_platform(platform, db_key, db_seller_name):
                continue
            
            sale_rows.append({
                'no': len(sale_rows) + 1,
                'invoice_no': invoice_no,
                'type': platform,
                'invoice_date': invoice_date,
                'warehouse_name': db_seller_name,
                'warehouse_code': warehouse_code,
                'gst_no': gst_no,
                'order_id': order_id,
                'item_asin': item_asin,
                'item_sku': item_sku,
                'item_name': item_name,
                'hsn_number': hsn_number,
                'quantity': db_qty,
                'item_cost': db_rate,
                'gross': gross_val,
                'igst': igst_val, 'cgst': cgst_val, 'sgst': sgst_val,
                'igst_amt': igst_amt_val, 'cgst_amt': cgst_amt_val, 'sgst_amt': sgst_amt_val,
                'invoice': invoice_val,
                'reason': reason,
                'zoho_status': zoho_status,
                'invoice_id': invoice_id,
                'state_code': db_state,
                'sale_unique_id': db_key,
                'calc_qty': db_qty,
                'calc_cost': db_rate,
                'calc_gross': round(db_qty * db_rate, 2),
                'calc_igst': 0, 'calc_cgst': 0, 'calc_sgst': 0,
                'calc_igst_amt': 0, 'calc_cgst_amt': 0, 'calc_sgst_amt': 0,
                'calc_invoice': round(db_qty * db_rate, 2),
                'state_code_short': f"({db_state})" if db_state else "",
                'sheet_name': db_sheet
            })
            
        purchase_rows = []
        for row in purchase_results:
            db_key, db_qty, db_rate, db_seller_name, db_sheet = row[0:5]
            (
                invoice_no, warehouse_code, gst_no, order_id,
                item_asin, item_sku, item_name, hsn_number,
                invoice_val,
                gross_val, igst_val, cgst_val, sgst_val, igst_amt_val, cgst_amt_val, sgst_amt_val
            ) = row[5:]
            
            if not matches_platform(platform, db_key, db_seller_name):
                continue
                    
            purchase_rows.append({
                'no': len(purchase_rows) + 1,
                'invoice_no': invoice_no,
                'warehouse_name': db_seller_name,
                'warehouse_code': warehouse_code,
                'gst_no': gst_no,
                'order_id': order_id,
                'item_asin': item_asin,
                'item_sku': item_sku,
                'item_name': item_name,
                'hsn_number': hsn_number,
                'quantity': db_qty,
                'item_cost': db_rate,
                'gross': gross_val,
                'igst': igst_val, 'cgst': cgst_val, 'sgst': sgst_val,
                'igst_amt': igst_amt_val, 'cgst_amt': cgst_amt_val, 'sgst_amt': sgst_amt_val,
                'invoice': invoice_val,
                'purchase_unique_id': db_key,
                'calc_qty': db_qty,
                'calc_cost': db_rate,
                'calc_gross': round(db_qty * db_rate, 2),
                'calc_igst': 0, 'calc_cgst': 0, 'calc_sgst': 0,
                'calc_igst_amt': 0, 'calc_cgst_amt': 0, 'calc_sgst_amt': 0,
                'calc_total_amt': round(db_qty * db_rate, 2),
                'sheet_name': db_sheet
            })

        # Merge with locally uploaded data
        try:
            stored = load_stored_data()
            local_sales = stored.get('sale', [])
            local_purchases = stored.get('purchase', [])
            
            for row in local_sales:
                db_sheet = row.get('sheet_name', get_financial_year())
                if year != 'ALL' and year not in db_sheet:
                    continue
                if not matches_platform(platform, row.get('sale_unique_id', ''), row.get('warehouse_name', ''), row.get('type', '')):
                    continue
                row['sheet_name'] = db_sheet
                row['no'] = len(sale_rows) + 1
                sale_rows.append(row)
                
            for row in local_purchases:
                db_sheet = row.get('sheet_name', get_financial_year())
                if year != 'ALL' and year not in db_sheet:
                    continue
                if not matches_platform(platform, row.get('purchase_unique_id', ''), row.get('warehouse_name', '')):
                    continue
                row['sheet_name'] = db_sheet
                row['no'] = len(purchase_rows) + 1
                purchase_rows.append(row)
        except Exception:
            pass

        # Organize by sheet/year dynamically based on matched sheets
        sheets_data = {}
        
        # Helper to extract year suffix
        import re
        def get_year_key(sh):
            m = re.search(r'\d{4}-\d{2}', sh)
            return m.group(0) if m else sh

        # Find all distinct years in results to construct keys
        found_years = set()
        for r in sale_rows:
            found_years.add(get_year_key(r.get('sheet_name', '')))
        for r in purchase_rows:
            found_years.add(get_year_key(r.get('sheet_name', '')))
            
        for y_key in found_years:
            if y_key:
                sheets_data[y_key] = { 'sale': [], 'purchase': [] }
                
        # If no years found, make sure we have current year as default
        if not sheets_data:
            sheets_data[get_financial_year()] = { 'sale': [], 'purchase': [] }
            
        for row in sale_rows:
            y_key = get_year_key(row.get('sheet_name', ''))
            if y_key in sheets_data:
                sheets_data[y_key]['sale'].append(row)
            else:
                # Default fallback
                first_key = list(sheets_data.keys())[0]
                sheets_data[first_key]['sale'].append(row)
                
        for row in purchase_rows:
            y_key = get_year_key(row.get('sheet_name', ''))
            if y_key in sheets_data:
                sheets_data[y_key]['purchase'].append(row)
            else:
                first_key = list(sheets_data.keys())[0]
                sheets_data[first_key]['purchase'].append(row)
                
        # Re-index row 'no' within each sheet
        for yr in sheets_data:
            for i, row in enumerate(sheets_data[yr]['sale']):
                row['no'] = i + 1
            for i, row in enumerate(sheets_data[yr]['purchase']):
                row['no'] = i + 1

        response = jsonify({
            "status": "Success",
            "data": sheets_data
        })
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response
    except Exception as e:
        traceback.print_exc()
        response = jsonify({"status": "Error", "message": str(e)})
        response.headers.add("Access-Control-Allow-Origin", "*")
        return response, 500


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads, process data, merge with existing, deduplicate."""
    try:
        required_files = ['sale_details', 'sale_summary', 'purchase_details', 'purchase_summary']
        file_paths = {}
        
        for file_key in required_files:
            if file_key not in request.files:
                return jsonify({'error': f'Missing file: {file_key}'}), 400
            
            file = request.files[file_key]
            if file.filename == '':
                return jsonify({'error': f'No file selected for: {file_key}'}), 400
            
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'{file_key}_{file.filename}')
            file.save(filepath)
            file_paths[file_key] = filepath
        
        # Process new uploaded files
        new_sale_data, new_purchase_data = process_files(
            file_paths['sale_details'],
            file_paths['sale_summary'],
            file_paths['purchase_details'],
            file_paths['purchase_summary']
        )
        
        # Load existing stored data
        stored = load_stored_data()
        
        # Merge with deduplication using UNIQUE IDs
        merged_sale, sale_ids, sale_new, sale_dups = merge_with_dedup(
            stored['sale'], new_sale_data,
            stored['sale_ids'], 'sale_unique_id'
        )
        
        merged_purchase, purchase_ids, purchase_new, purchase_dups = merge_with_dedup(
            stored['purchase'], new_purchase_data,
            stored['purchase_ids'], 'purchase_unique_id'
        )
        
        # Save merged data persistently
        stored_data = {
            'sale': merged_sale,
            'purchase': merged_purchase,
            'sale_ids': sale_ids,
            'purchase_ids': purchase_ids,
        }
        save_stored_data(stored_data)
        

        # Return preview (first 100 rows)
        preview_sale = merged_sale[:100]
        preview_purchase = merged_purchase[:100]
        
        return jsonify({
            'success': True,
            'sale_count': len(merged_sale),
            'purchase_count': len(merged_purchase),
            'sale_new': sale_new,
            'sale_duplicates': sale_dups,
            'purchase_new': purchase_new,
            'purchase_duplicates': purchase_dups,
            'financial_year': get_financial_year(),
            'sale_preview': preview_sale,
            'purchase_preview': preview_purchase,
        })
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/export', methods=['GET'])
def export_file():
    """Export processed data as Excel."""
    try:
        # Try persistent store first, then temp
        stored = load_stored_data()
        if stored['sale'] or stored['purchase']:
            sale_data = stored['sale']
            purchase_data = stored['purchase']
        else:
            data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'processed_data.pkl')
            if not os.path.exists(data_path):
                return jsonify({'error': 'No processed data. Please upload files first.'}), 400
            with open(data_path, 'rb') as f:
                sale_data, purchase_data = pickle.load(f)
        
        fy = get_financial_year()
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'AJIO_MYNTRA_FLIPKART_{fy}_Output.xlsx')
        export_to_excel(sale_data, purchase_data, output_path)
        
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'AJIO_MYNTRA_FLIPKART_Sale_Purchase_{fy}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/preview-more', methods=['GET'])
def preview_more():
    """Get more rows for preview."""
    try:
        # Try persistent store first
        stored = load_stored_data()
        if stored['sale'] or stored['purchase']:
            sale_data = stored['sale']
            purchase_data = stored['purchase']
        else:
            data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'processed_data.pkl')
            if not os.path.exists(data_path):
                return jsonify({'error': 'No data available'}), 400
            with open(data_path, 'rb') as f:
                sale_data, purchase_data = pickle.load(f)
        
        offset = int(request.args.get('offset', 0))
        limit = int(request.args.get('limit', 100))
        
        return jsonify({
            'sale_data': sale_data[offset:offset + limit],
            'purchase_data': purchase_data[offset:offset + limit],
            'sale_total': len(sale_data),
            'purchase_total': len(purchase_data),
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/status', methods=['GET'])
def status():
    """Get current data status."""
    stored = load_stored_data()
    return jsonify({
        'sale_count': len(stored['sale']),
        'purchase_count': len(stored['purchase']),
        'financial_year': get_financial_year(),
        'has_data': bool(stored['sale'] or stored['purchase']),
    })


@app.route('/clear', methods=['POST'])
def clear_data():
    """Clear all stored data (fresh start)."""
    try:
        empty_store = {'sale': [], 'purchase': [], 'sale_ids': set(), 'purchase_ids': set()}
        save_stored_data(empty_store)

        data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'processed_data.pkl')
        try:
            if os.path.exists(data_path):
                os.remove(data_path)
        except OSError:
            with open(data_path, 'wb') as f:
                pickle.dump(([], []), f)

        return jsonify({'success': True, 'message': 'All data cleared!'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/get-sync-data', methods=['POST'])
def get_sync_data():
    """Get processed and deduplicated data prepared for Google Sheets sync."""
    try:
        req_data = request.json or {}
        web_app_url = req_data.get('webAppUrl')
        sheet_name = req_data.get('sheetName')
        mode = req_data.get('mode', 'append')
        
        if not web_app_url:
            return jsonify({'error': 'Google Web App URL is required'}), 400
        if not sheet_name:
            return jsonify({'error': 'Target Sheet Name is required'}), 400
            
        stored = load_stored_data()
        sale_data = [dict(row) for row in stored['sale']]
        purchase_data = [dict(row) for row in stored['purchase']]
        
        if not sale_data and not purchase_data:
            return jsonify({'error': 'No data processed yet. Please upload files first.'}), 400
            
        import requests
        from urllib.parse import quote
        
        original_sale_count = len(sale_data)
        original_purchase_count = len(purchase_data)
        skipped_sale = 0
        skipped_purchase = 0
        
        if mode == 'append':
            try:
                url = f"{web_app_url}?action=getExistingUniqueIds&sheetName={quote(sheet_name)}"
                res = requests.get(url, allow_redirects=False)
                if res.status_code in [302, 307, 308]:
                    redirect_url = res.headers.get('Location')
                    res = requests.get(redirect_url)
                
                if res.status_code == 200:
                    res_data = res.json()
                    if res_data.get('success'):
                        existing_sale_ids = set(str(uid) for uid in res_data.get('saleIds', []))
                        existing_purchase_ids = set(str(uid) for uid in res_data.get('purchaseIds', []))
                        
                        sale_filtered = [row for row in sale_data if str(row.get('sale_unique_id', '')) not in existing_sale_ids]
                        purchase_filtered = [row for row in purchase_data if str(row.get('purchase_unique_id', '')) not in existing_purchase_ids]
                        
                        skipped_sale = len(sale_data) - len(sale_filtered)
                        skipped_purchase = len(purchase_data) - len(purchase_filtered)
                        
                        sale_data = sale_filtered
                        purchase_data = purchase_filtered
                        
                        for idx, row in enumerate(sale_data):
                            row['no'] = len(existing_sale_ids) + idx + 1
                        for idx, row in enumerate(purchase_data):
                            row['no'] = len(existing_purchase_ids) + idx + 1
            except Exception as e:
                print(f"Error fetching unique IDs: {str(e)}")
                
        # Map dict lists to 2D lists matching Google Sheet column layouts
        sale_keys = [
            'no', 'invoice_no', 'type', 'invoice_date', 'warehouse_name',
            'warehouse_code', 'gst_no', 'order_id', 'item_asin', 'item_sku',
            'item_name', 'hsn_number', 'quantity', 'item_cost', 'gross',
            'igst', 'cgst', 'sgst', 'igst_amt', 'cgst_amt', 'sgst_amt',
            'invoice', 'reason', 'zoho_status', 'invoice_id', 'state_code',
            'sale_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
            'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
            'calc_cgst_amt', 'calc_sgst_amt', 'calc_invoice', 'state_code_short'
        ]
        
        purchase_keys = [
            'no', 'invoice_no', 'warehouse_name', 'warehouse_code', 'gst_no',
            'order_id', 'item_asin', 'item_sku', 'item_name', 'hsn_number',
            'quantity', 'item_cost', 'gross', 'igst', 'cgst', 'sgst',
            'igst_amt', 'cgst_amt', 'sgst_amt', 'invoice',
            'purchase_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
            'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
            'calc_cgst_amt', 'calc_sgst_amt', 'calc_total_amt'
        ]
        
        sale_rows_2d = [[row.get(k, '') for k in sale_keys] for row in sale_data]
        purchase_rows_2d = [[row.get(k, '') for k in purchase_keys] for row in purchase_data]
        
        return jsonify({
            'success': True,
            'saleRows': sale_rows_2d,
            'purchaseRows': purchase_rows_2d,
            'originalSale': original_sale_count,
            'originalPurchase': original_purchase_count,
            'skippedSale': skipped_sale,
            'skippedPurchase': skipped_purchase,
            'toSyncSale': len(sale_rows_2d),
            'toSyncPurchase': len(purchase_rows_2d),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/sync-sheet', methods=['POST'])
def sync_sheet():
    """Sync stored data to Google Sheet via the Google Web App URL."""
    try:
        req_data = request.json or {}
        web_app_url = req_data.get('webAppUrl')
        sheet_name = req_data.get('sheetName')
        mode = req_data.get('mode', 'append')  # 'append' or 'overwrite'
        
        if not web_app_url:
            return jsonify({'error': 'Google Web App URL is required'}), 400
        if not sheet_name:
            return jsonify({'error': 'Target Sheet Name is required'}), 400
            
        # Load the stored data
        stored = load_stored_data()
        sale_data = [dict(row) for row in stored['sale']]
        purchase_data = [dict(row) for row in stored['purchase']]
        
        if not sale_data and not purchase_data:
            return jsonify({'error': 'No data processed yet. Please upload files first.'}), 400

        import requests
        from urllib.parse import quote

        # If in append mode, fetch existing unique IDs to avoid duplication in Google Sheets
        if mode == 'append':
            try:
                # Call getExistingUniqueIds from Google Sheets Web App
                url = f"{web_app_url}?action=getExistingUniqueIds&sheetName={quote(sheet_name)}"
                res = requests.get(url, allow_redirects=False)
                if res.status_code in [302, 307, 308]:
                    redirect_url = res.headers.get('Location')
                    res = requests.get(redirect_url)
                
                if res.status_code == 200:
                    res_data = res.json()
                    if res_data.get('success'):
                        existing_sale_ids = set(str(uid) for uid in res_data.get('saleIds', []))
                        existing_purchase_ids = set(str(uid) for uid in res_data.get('purchaseIds', []))
                        
                        # Filter out rows that already exist in the sheet
                        sale_data = [row for row in sale_data if str(row.get('sale_unique_id', '')) not in existing_sale_ids]
                        purchase_data = [row for row in purchase_data if str(row.get('purchase_unique_id', '')) not in existing_purchase_ids]
                        
                        # Re-number the appended rows starting after the existing counts
                        for idx, row in enumerate(sale_data):
                            row['no'] = len(existing_sale_ids) + idx + 1
                        for idx, row in enumerate(purchase_data):
                            row['no'] = len(existing_purchase_ids) + idx + 1
            except Exception as e:
                # Log the error but proceed with default append if it fails
                print(f"Error fetching unique IDs for append deduplication: {str(e)}")
            
        # Map dict lists to 2D lists matching Google Sheet column layouts
        sale_keys = [
            'no', 'invoice_no', 'type', 'invoice_date', 'warehouse_name',
            'warehouse_code', 'gst_no', 'order_id', 'item_asin', 'item_sku',
            'item_name', 'hsn_number', 'quantity', 'item_cost', 'gross',
            'igst', 'cgst', 'sgst', 'igst_amt', 'cgst_amt', 'sgst_amt',
            'invoice', 'reason', 'zoho_status', 'invoice_id', 'state_code',
            'sale_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
            'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
            'calc_cgst_amt', 'calc_sgst_amt', 'calc_invoice', 'state_code_short'
        ]
        
        purchase_keys = [
            'no', 'invoice_no', 'warehouse_name', 'warehouse_code', 'gst_no',
            'order_id', 'item_asin', 'item_sku', 'item_name', 'hsn_number',
            'quantity', 'item_cost', 'gross', 'igst', 'cgst', 'sgst',
            'igst_amt', 'cgst_amt', 'sgst_amt', 'invoice',
            'purchase_unique_id', 'calc_qty', 'calc_cost', 'calc_gross',
            'calc_igst', 'calc_cgst', 'calc_sgst', 'calc_igst_amt',
            'calc_cgst_amt', 'calc_sgst_amt', 'calc_total_amt'
        ]
        
        sale_rows_2d = [[row.get(k, '') for k in sale_keys] for row in sale_data]
        purchase_rows_2d = [[row.get(k, '') for k in purchase_keys] for row in purchase_data]
        
        # Prepare the payload for Google Apps Script Web App
        payload = {
            'action': 'writeData',
            'sheetName': sheet_name,
            'saleRows': sale_rows_2d,
            'purchaseRows': purchase_rows_2d,
            'isAppend': (mode == 'append')
        }
        
        # Call the Google Web App
        import requests
        # Manually handle 302 redirects to preserve POST method
        response = requests.post(web_app_url, json=payload, allow_redirects=False)
        if response.status_code in [302, 307, 308]:
            redirect_url = response.headers.get('Location')
            response = requests.post(redirect_url, json=payload)
            
        if response.status_code != 200:
            return jsonify({'error': f'Failed to write to Google Sheets. Status code: {response.status_code}'}), 500
            
        result = response.json()
        if not result.get('success'):
            return jsonify({'error': result.get('error', 'Unknown error writing to Google Sheets')}), 500
            
        return jsonify({
            'success': True,
            'message': result.get('message', 'Data synced successfully!'),
            'saleRows': result.get('saleRows', 0),
            'purchaseRows': result.get('purchaseRows', 0)
        })
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, use_reloader=False, threaded=True, port=port)
